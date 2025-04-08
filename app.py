from flask import Flask, request, jsonify, send_file
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
ALLOWED_EXTENSIONS = {'sb3'}
EXCEL_FILE = 'submissions.xlsx'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['RESULT_FOLDER'] = RESULT_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def write_to_excel(name, student_class, score):
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Submissions"
        sheet.append(["Name", "Class", "Score"])
    else:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

    sheet.append([name, student_class, score])
    workbook.save(EXCEL_FILE)

def generate_result_file(name, student_class, score):
    filename = f"{name.replace(' ', '_')}_result.txt"
    filepath = os.path.join(RESULT_FOLDER, filename)
    with open(filepath, 'w') as f:
        f.write(f"🎓 Name: {name}\n")
        f.write(f"🏫 Class: {student_class}\n")
        f.write(f"✅ Score: {score}/10\n")
        f.write("\nThank you for taking the PictoBlox AI test!")
    return filepath

@app.route('/submit_quiz', methods=['POST'])
def submit_quiz():
    name = request.form.get("name")
    student_class = request.form.get("class")

    if 'projectFile' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['projectFile']

    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file and allowed_file(file.filename):
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)

        score = 8  # Replace this with actual score logic
        write_to_excel(name, student_class, score)

        result_path = generate_result_file(name, student_class, score)

        return jsonify({'message': 'File uploaded and data saved', 'score': score, 'download_url': f"/download_result/{os.path.basename(result_path)}"})
    else:
        return jsonify({'error': 'Invalid file format'}), 400

@app.route('/download_result/<filename>', methods=['GET'])
def download_result(filename):
    return send_file(os.path.join(app.config['RESULT_FOLDER'], filename), as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
